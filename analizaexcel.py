from openpyxl import load_workbook
 
def analizar_excel(ruta_archivo):
    try:
        wb = load_workbook(ruta_archivo)
        print(f"Analizando: {ruta_archivo}")
 
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- Hoja: {sheet_name} ---")
 
            # Contar imágenes
            num_imagenes = len(ws._images)
            if num_imagenes > 0:
                print(f"  Contiene {num_imagenes} imágenes.")
 
            # Contar gráficos
            num_graficos = len(ws._charts)
            if num_graficos > 0:
                print(f"  Contiene {num_graficos} gráficos.")
 
            # Contar formas y otros objetos (requiere más detalle, ya que openpyxl no los categoriza tan directamente como imágenes/gráficos)
            # Para formas, necesitarías inspeccionar 'drawing' o 'vml' si quieres profundizar.
            # Esta es una aproximación para ver si hay "dibujos" generales:
            if ws.drawing is not None and len(ws.drawing._shapes) > 0:
                print(f"  Contiene formas o elementos de dibujo.")
 
        print("\nAnálisis completado.")
 
    except Exception as e:
        print(f"Error al analizar el archivo: {e}")
 
# Ejemplo de uso:
analizar_excel("C:/Users/David Dimas/Downloads/Doc 302-303 update.xlsx")